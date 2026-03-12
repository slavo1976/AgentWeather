"""
WeatherAgent — stiahne 7-dňovú predpoveď z yr.no (MET Norway API)
pre 3 mestá, uloží do WeatherHistory.xlsx a pushne na GitHub.

Požiadavky:
  pip install requests openpyxl PyGithub

Env premenné:
  GITHUB_TOKEN   — GitHub Personal Access Token (scope: repo)
  GITHUB_USER    — tvoje GitHub používateľské meno
"""

import os
import io
import json
import math
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone, date, timedelta
from collections import defaultdict
from github import Github, Auth, GithubException

# ── Konfigurácia ──────────────────────────────────────────────────────────────

CITIES = [
    {"name": "Leonidio",         "lat": 37.1547,  "lon": 22.8742,  "country": "Greece"},
    {"name": "Arco",             "lat": 45.9167,  "lon": 10.8833,  "country": "Italy"},
    {"name": "San Vito la Capo", "lat": 38.1739,  "lon": 12.7340,  "country": "Sicily, Italy"},
    {"name": "Margalef",         "lat": 41.2833,  "lon":  0.8167,  "country": "Spain"},
    {"name": "Barsnes",          "lat": 61.1833,  "lon":  6.7667,  "country": "Norway"},
]

REPO_NAME   = "AgentWeather"
EXCEL_FILE  = "WeatherHistory.xlsx"
HEADERS = ["Dátum", "Teplota (°C)", "Množstvo zrážok (mm)",
           "Rýchlosť vetra (m/s)", "Smer vetra", "Oblačnosť (%)", "Dĺžka slnečného svitu (h)"]

# ── Pomocné funkcie ───────────────────────────────────────────────────────────

def wind_degrees_to_text(deg: float | None) -> str:
    if deg is None:
        return ""
    dirs = ["S", "S", "SZ", "SZ", "Z", "Z", "JZ", "JZ",
            "J", "J", "JV", "JV", "V", "V", "SV", "SV"]
    return dirs[round(deg / 22.5) % 16]


def fetch_forecast(city: dict) -> list[dict]:
    """
    Stiahne hodinové dáta z MET Norway API a agreguje ich po dňoch
    na 7-dňovú predpoveď. Vráti zoznam dict pre každý deň.
    """
    url = "https://api.met.no/weatherapi/locationforecast/2.0/compact"
    headers = {"User-Agent": "WeatherAgent/1.0 github.com/AgentWeather"}
    params  = {"lat": city["lat"], "lon": city["lon"]}

    resp = requests.get(url, headers=headers, params=params, timeout=20)
    resp.raise_for_status()
    data = resp.json()

    # Zbieraj hodinové záznamy podľa dátumu
    day_buckets: dict[date, list] = defaultdict(list)
    today = datetime.now(timezone.utc).date()

    for entry in data["properties"]["timeseries"]:
        dt    = datetime.fromisoformat(entry["time"].replace("Z", "+00:00"))
        d     = dt.date()
        delta = (d - today).days
        if delta < 0 or delta >= 7:
            continue

        instant = entry["data"].get("instant", {}).get("details", {})
        next1   = entry["data"].get("next_1_hours",  {}).get("details", {})
        next6   = entry["data"].get("next_6_hours",  {}).get("details", {})

        precip = (
            next1.get("precipitation_amount") or
            next6.get("precipitation_amount")
        )
        day_buckets[d].append({
            "temp":      instant.get("air_temperature"),
            "wind_spd":  instant.get("wind_speed"),
            "wind_dir":  instant.get("wind_from_direction"),
            "clouds":    instant.get("cloud_area_fraction"),
            "precip":    precip,
        })

    # Agregácia na deň (priemer / suma)
    result = []
    for d in sorted(day_buckets)[:7]:
        rows = day_buckets[d]

        def avg(key):
            vals = [r[key] for r in rows if r[key] is not None]
            return round(sum(vals) / len(vals), 1) if vals else None

        def total(key):
            vals = [r[key] for r in rows if r[key] is not None]
            return round(sum(vals), 1) if vals else None

        cloud_avg = avg("clouds")
        # Odhad dĺžky slnečného svitu: max 12h × (1 − oblačnosť/100)
        sunshine = round(12 * (1 - (cloud_avg or 0) / 100), 1) if cloud_avg is not None else None

        result.append({
            "date":     d,
            "temp":     avg("temp"),
            "precip":   total("precip"),
            "wind_spd": avg("wind_spd"),
            "wind_dir": wind_degrees_to_text(avg("wind_dir")),
            "clouds":   cloud_avg,
            "sunshine": sunshine,
        })

    return result


# ── Excel operácie ────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DATA_FONT   = Font(name="Arial", size=10)
THIN        = Side(style="thin", color="BFBFBF")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALT_FILL    = PatternFill("solid", start_color="EBF3FB")


def style_sheet(ws, city_info: dict):
    # Riadok 1: Google Maps link
    maps_url = f"https://www.google.com/maps?q={city_info['lat']},{city_info['lon']}"
    label    = f"📍 {city_info['name']}, {city_info['country']} — Google Maps"
    link_cell = ws.cell(row=1, column=1, value=label)
    link_cell.hyperlink  = maps_url
    link_cell.font       = Font(bold=True, color="0563C1", name="Arial", size=10, underline="single")
    link_cell.alignment  = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws.row_dimensions[1].height = 20

    # Riadok 2: hlavičky stĺpcov
    ws.freeze_panes = "B3"
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font       = HEADER_FONT
        cell.fill       = HEADER_FILL
        cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border     = BORDER

    col_widths = [14, 14, 20, 20, 14, 14, 22]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 32


def upsert_sheet(ws, forecast: list[dict]):
    """Aktualizuje existujúce riadky alebo pridáva nové."""
    # Načítaj existujúce dátumy → riadok
    date_to_row: dict[date, int] = {}
    for row_idx in range(3, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=1).value
        if isinstance(cell_val, (datetime, date)):
            d = cell_val.date() if isinstance(cell_val, datetime) else cell_val
            date_to_row[d] = row_idx

    for rec in forecast:
        d       = rec["date"]
        row_idx = date_to_row.get(d)

        if row_idx is None:
            # Nový riadok (min. riadok 3 — riadok 1 je link, riadok 2 je hlavička)
            row_idx = max(ws.max_row + 1, 3)
            date_to_row[d] = row_idx

        values = [
            d,
            rec["temp"],
            rec["precip"],
            rec["wind_spd"],
            rec["wind_dir"],
            rec["clouds"],
            rec["sunshine"],
        ]
        alt = (row_idx % 2 == 0)
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font   = DATA_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
            if alt:
                cell.fill = ALT_FILL
            if col_idx == 1 and isinstance(val, date):
                cell.number_format = "DD.MM.YYYY"


def build_workbook(all_forecasts: dict[str, list[dict]],
                   existing_wb: openpyxl.Workbook | None = None) -> openpyxl.Workbook:
    wb = existing_wb or openpyxl.Workbook()

    # Odstráň predvolený list ak nový
    if existing_wb is None and "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for city_info in CITIES:
        city_name = city_info["name"]
        forecast  = all_forecasts[city_name]

        if city_name in wb.sheetnames:
            ws = wb[city_name]
        else:
            ws = wb.create_sheet(city_name)
            ws.title = city_name
            style_sheet(ws, city_info)

        upsert_sheet(ws, forecast)

    # Uisti sa, že prvý sheet je aktívny
    wb.active = wb.worksheets[0]
    return wb


# ── GitHub operácie ───────────────────────────────────────────────────────────

def push_to_github(wb: openpyxl.Workbook, token: str, github_user: str) -> str:
    """
    Pushne WeatherHistory.xlsx do repozitára AgentWeather.
    Ak repo neexistuje, vytvorí ho. Vráti URL súboru.
    """
    g    = Github(auth=Auth.Token(token))
    user = g.get_user()

    # Získaj alebo vytvor repo
    try:
        repo = user.get_repo(REPO_NAME)
        print(f"  Repozitár '{REPO_NAME}' nájdený.")
    except GithubException:
        repo = user.create_repo(
            REPO_NAME,
            description="Automatická 7-dňová predpoveď počasia",
            private=False,
            auto_init=True,
        )
        print(f"  Repozitár '{REPO_NAME}' vytvorený.")

    # Serializuj workbook do bytov
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    commit_msg = f"WeatherHistory update – {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}"

    # Aktualizuj alebo vytvor súbor
    try:
        existing = repo.get_contents(EXCEL_FILE)
        repo.update_file(EXCEL_FILE, commit_msg, content, existing.sha)
        print(f"  Súbor '{EXCEL_FILE}' aktualizovaný.")
    except GithubException:
        repo.create_file(EXCEL_FILE, commit_msg, content)
        print(f"  Súbor '{EXCEL_FILE}' vytvorený.")

    return f"https://github.com/{user.login}/{REPO_NAME}/blob/main/{EXCEL_FILE}"


# ── Hlavná funkcia ────────────────────────────────────────────────────────────

def main():
    token  = os.environ.get("GITHUB_TOKEN")
    g_user = os.environ.get("GITHUB_USER", "")

    if not token:
        raise EnvironmentError(
            "Nastav GITHUB_TOKEN:\n"
            "  export GITHUB_TOKEN='ghp_...'\n"
            "  export GITHUB_USER='tvoje-meno'  # voliteľné"
        )

    print("=== WeatherAgent štartuje ===\n")

    # 1. Stiahni predpovede
    all_forecasts: dict[str, list[dict]] = {}
    for city in CITIES:
        print(f"Sťahujem predpoveď pre {city['name']} ({city['country']})...")
        all_forecasts[city["name"]] = fetch_forecast(city)
        print(f"  → {len(all_forecasts[city['name']])} dní stiahnutých.")

    # 2. Zostav workbook (pokusom načítaj existujúci zo GitHub)
    existing_wb = None
    try:
        g    = Github(auth=Auth.Token(token))
        repo = g.get_user().get_repo(REPO_NAME)
        file = repo.get_contents(EXCEL_FILE)
        buf  = io.BytesIO(file.decoded_content)
        existing_wb = openpyxl.load_workbook(buf)
        print(f"\nExistujúci '{EXCEL_FILE}' načítaný z GitHub (upsert móde).")
    except Exception:
        print(f"\nExistujúci '{EXCEL_FILE}' nenájdený — vytváram nový.")

    print("\nBuduje sa Excel workbook...")
    wb = build_workbook(all_forecasts, existing_wb)

    # Uloži lokálnu kópiu
    local_path = f"/tmp/{EXCEL_FILE}"
    wb.save(local_path)
    print(f"  Lokálna kópia uložená: {local_path}")

    # 3. Push na GitHub
    print("\nNahráva sa na GitHub...")
    url = push_to_github(wb, token, g_user)

    print(f"\n✅ Hotovo!\nGitHub link: {url}")
    return url


if __name__ == "__main__":
    main()
